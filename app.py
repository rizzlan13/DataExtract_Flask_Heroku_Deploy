from flask import Flask, request, render_template, jsonify, redirect, url_for
import os
import re
import fitz
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from tqdm import tqdm  # Import tqdm for progress bar
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

# Global variables to store bounding box coordinates, extracted text, and PDF filenames
bounding_boxes = []
extracted_texts = []
pdf_filenames = []
bounding_boxes_drawn = False

app = Flask(__name__)

# Your existing functions
# Define your functions here or import them from another module

# Function to get PDF file paths from a folder and its subdirectories
def get_pdf_paths(folder_path):
    pdf_paths = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith('.pdf'):
                pdf_paths.append(os.path.join(root, file))
    return pdf_paths

def draw_bounding_boxes_single(pdf_path):
    global bounding_boxes, pdf_filenames, extracted_texts  # Use the global variables
    bounding_boxes = []  # Reset bounding boxes for a new PDF

    # Extract the filename with .pdf extension from the PDF file path
    pdf_filename = os.path.basename(pdf_path)
    pdf_filenames.append(pdf_filename)

    # Load the PDF
    pdf_document = fitz.open(pdf_path)
    page = pdf_document.load_page(0)  # Load the first page (you can loop through pages)

    # Convert PDF page to an image
    image = page.get_pixmap()
    img = Image.frombytes("RGB", [image.width, image.height], image.samples)

    # Initialize a tkinter window
    window = tk.Tk()
    window.title("PDF Bounding Box Extractor")

    # Modify the label creation and configuration to set font color and style
    label = tk.Label(window, text="", font=("Arial", 12, "bold"), fg="green")
    label.pack()

    # Variables to store mouse drag coordinates
    start_x, start_y = None, None
    rect_id = None

    # Function to handle mouse button press event
    def on_press(event):
        nonlocal start_x, start_y, rect_id
        start_x, start_y = event.x, event.y

    # Function to handle mouse drag event and draw rectangle
    def on_drag(event):
        nonlocal start_x, start_y, rect_id
        if rect_id:
            canvas.delete(rect_id)  # Delete previous rectangle
        rect_id = canvas.create_rectangle(start_x, start_y, event.x, event.y, outline='red', width=2)

    # Function to handle mouse button release event and store bounding box
    def on_release(event):
        nonlocal start_x, start_y, rect_id
        if rect_id:
            canvas.delete(rect_id)  # Delete the rectangle
        end_x, end_y = event.x, event.y
        bbox = (start_x, start_y, end_x, end_y)
        bounding_boxes.append(bbox)
        canvas.create_rectangle(*bbox, outline='red', width=2)  # Draw the final rectangle

        # Extract data based on bounding boxes
        extracted_data = extract_data(pdf_path, bounding_boxes)
        # cleaned_data = clean_extracted_text(extracted_data)
        extracted_texts.append(extracted_data)
        label.config(text=f"The data for the file named {pdf_filename} was successfully extracted.")

        # Automatically close the window and proceed to the next PDF
        window.after(250, window.destroy)  # Close after a delay (in milliseconds)

    # Convert PDF page to an image and display
    image = page.get_pixmap()
    img_tk = ImageTk.PhotoImage(Image.frombytes("RGB", [image.width, image.height], image.samples))

    canvas = tk.Canvas(window, width=img_tk.width(), height=img_tk.height())
    canvas.pack()
    canvas.create_image(0, 0, anchor=tk.NW, image=img_tk)

    # Bind mouse events to their respective functions
    canvas.bind("<ButtonPress-1>", on_press)
    canvas.bind("<B1-Motion>", on_drag)
    canvas.bind("<ButtonRelease-1>", on_release)

    window.mainloop()

def draw_bounding_boxes_first(pdf_path):
    global bounding_boxes, pdf_filenames, extracted_texts, bounding_boxes_drawn  # Use the global variables

    if not bounding_boxes_drawn:
        bounding_boxes = []  # Reset bounding boxes for a new PDF

        # Extract the filename with .pdf extension from the PDF file path
        pdf_filename = os.path.basename(pdf_path)
        pdf_filenames.append(pdf_filename)

        # Load the PDF
        pdf_document = fitz.open(pdf_path)
        page = pdf_document.load_page(0)  # Load the first page (you can loop through pages)

        # Convert PDF page to an image
        image = page.get_pixmap()
        img = Image.frombytes("RGB", [image.width, image.height], image.samples)

        # Initialize a tkinter window
        window = tk.Tk()
        window.title("PDF Bounding Box Extractor")

        # Modify the label creation and configuration to set font color and style
        label = tk.Label(window, text="", font=("Arial", 12, "bold"), fg="green")
        label.pack()

        # Variables to store mouse drag coordinates
        start_x, start_y = None, None
        rect_id = None

        # Function to handle mouse button press event
        def on_press(event):
            nonlocal start_x, start_y, rect_id
            start_x, start_y = event.x, event.y

        # Function to handle mouse drag event and draw rectangle
        def on_drag(event):
            nonlocal start_x, start_y, rect_id
            if rect_id:
                canvas.delete(rect_id)  # Delete previous rectangle
            rect_id = canvas.create_rectangle(start_x, start_y, event.x, event.y, outline='red', width=2)

        # Function to handle mouse button release event and store bounding box
        def on_release(event):
            nonlocal start_x, start_y, rect_id
            if rect_id:
                canvas.delete(rect_id)  # Delete the rectangle
            end_x, end_y = event.x, event.y
            bbox = (start_x, start_y, end_x, end_y)
            bounding_boxes.append(bbox)
            canvas.create_rectangle(*bbox, outline='red', width=2)  # Draw the final rectangle

            # Extract data based on bounding boxes
            # Within your existing code after text extraction
            extracted_data = extract_data(pdf_path, bounding_boxes)
            cleaned_data = clean_extracted_text(extracted_data)
            extracted_texts.append(cleaned_data)
            label.config(text=f"The data for the file named {pdf_filename} was successfully extracted.")

            # Automatically close the window and proceed to the next PDF
            window.after(250, window.destroy)  # Close after a delay (in milliseconds)

        # Convert PDF page to an image and display
        image = page.get_pixmap()
        img_tk = ImageTk.PhotoImage(Image.frombytes("RGB", [image.width, image.height], image.samples))

        canvas = tk.Canvas(window, width=img_tk.width(), height=img_tk.height())
        canvas.pack()
        canvas.create_image(0, 0, anchor=tk.NW, image=img_tk)

        # Bind mouse events to their respective functions
        canvas.bind("<ButtonPress-1>", on_press)
        canvas.bind("<B1-Motion>", on_drag)
        canvas.bind("<ButtonRelease-1>", on_release)

        window.mainloop()

def extract_data(pdf_path, bounding_boxes):
    pdf_document = fitz.open(pdf_path)
    page = pdf_document.load_page(0)

    extracted_data = []
    for bbox in bounding_boxes:
        x0, y0, x1, y1 = bbox
        
        words = page.get_text("words", clip=(x0, y0, x1, y1))
        words_method_text = " ".join(word[4] for word in words).strip()

        extracted_data.append(words_method_text)

    return extracted_data

def clean_extracted_text(extracted_text):
    cleaned_text = []
    for text in extracted_text:
        # Remove unwanted characters or patterns
        cleaned = re.sub(r"[^\w\s]", "", text)  # Remove non-alphanumeric characters except spaces
        cleaned = re.sub(r"\s+", " ", cleaned).strip()  # Remove extra spaces and strip leading/trailing spaces
        cleaned_text.append(cleaned)
    return cleaned_text

# Function to save extracted data to an Excel file
def save_to_excel(pdf_filenames, extracted_texts, folder_path):
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"{os.path.basename(folder_path)}_{now}.xlsx"

    ## ADDITIONAL COLUMN
    workbook = Workbook()
    main_sheet = workbook.active
    main_sheet.title = 'MainSheet'
    
    # Headers for MainSheet
    headers = ['FileName', 'Doc Type', 'Seq', 'No Sheet', 'Title', 'Rev', 'Year', 'Descp']
    main_sheet.append(headers)

    # Set header style: bold, center, middle align, and border
    header_style = Font(bold=True)
    align_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in main_sheet[1]:
        cell.font = header_style
        cell.alignment = align_style
        cell.border = border_style

    # Adding data to MainSheet without dropdowns for FileName and Title columns
    for idx, (pdf_file, extracted_data) in enumerate(zip(pdf_filenames, extracted_texts)):
        main_sheet.append([pdf_file] + extracted_data)  # Appending PDF filename and extracted data
        
    # Applying changes to Title column
    for row_idx, extracted_data_row in enumerate(extracted_texts, start=2):
        title_value = ' '.join(extracted_data_row)  # Combine extracted text data
        title_value = title_value.replace('\n', ' ').replace("'", "").replace('"', '').replace(':', '').replace(';', '').replace('_', '').replace('/', '').replace("&", "AND").replace("  ", " ")
        main_sheet.cell(row=row_idx, column=headers.index('Title') + 1).value = title_value

    # Applying changes to 'FileName' column (without dropdown)
    for row_idx, pdf_file in enumerate(pdf_filenames, start=2):
        main_sheet.cell(row=row_idx, column=headers.index('FileName') + 1).value = pdf_file
        
    # Apply formula to MainSheet column after 'Descp'
    desc_column_index = headers.index('Descp') + 1
    formula_col = desc_column_index + 1  # The column after 'Descp'

    for row in range(2, len(pdf_filenames) + 2):  # Start from the second row
        formula = f'=IFERROR(TRIM(RIGHT(E{row}, LEN(E{row}) - SEARCH(",", E{row}))), E{row})'  # Dynamic formula with auto-incrementing row number
        main_sheet.cell(row=row, column=formula_col).value = formula

    # Create SourceSheet and define data for SourceSheet
    source_sheet = workbook.create_sheet('SourceSheet')

    # Define data for SourceSheet
    doc_types = [
        'EML', 'FAX', 'LTR', 'MEM', 'MOU', 'TRM', 'VTRM', 'ABD', 'ACL', 'ARV', 'ASM', 'BFD', 'BOM', 'BOP', 'BPI', 'BR', 
        'CAA', 'CAL', 'CAR', 'CBD', 'CBE', 'CEM', 'CER', 'CHG', 'CKL', 'CKS', 'CLO', 'CLR', 'CO', 'COD', 'COM', 'COP', 
        'COS', 'CR', 'CRM', 'CSC', 'CSD', 'CTG', 'CTR', 'CUR', 'DBM', 'DC', 'DCN', 'DCR', 'DDR', 'DEQ', 'DES', 'DFS', 
        'DGM', 'DOS', 'DPR', 'DSP', 'DTL', 'DWG', 'EGA', 'EIA', 'ENV', 'EQD', 'EQL', 'ERP', 'ERT', 'EXD', 'FAL', 'FCN', 
        'FCR', 'FI', 'FID', 'FIR', 'FRM', 'FRR', 'FRT', 'FU', 'GA', 'GAR', 'HAC', 'HAZ', 'HDW', 'HEL', 'HMB', 'HOK', 
        'HRD', 'HUC', 'HYD', 'HZA', 'IAA', 'IDX', 'IND', 'INT', 'INV', 'ISO', 'ITP', 'ITR', 'JCT', 'JHA', 'JMS', 'JSA', 
        'JSP', 'LIC', 'LOG', 'LOP', 'LST', 'LYT', 'MAF', 'MAN', 'MDR', 'MGP', 'MOM', 'MPI', 'MPR', 'MPS', 'MQP', 'MRE', 
        'MRIR', 'MST', 'MTD', 'MTO', 'MVR', 'MVS', 'NCR', 'NDR', 'NOI', 'OCS', 'OMM', 'OPM', 'ORG', 'PCH', 'PCL', 'PDR', 
        'PEP', 'PFD', 'PFS', 'PID', 'PLL', 'PLN', 'PMI', 'PO', 'PPS', 'PR', 'PRC', 'PRP', 'PRT', 'PS', 'PSD', 'PSFD', 
        'PSP', 'PTV', 'PWT', 'QAT', 'QLF', 'QTD', 'RA', 'RCA', 'RDN', 'REG', 'RGT', 'RLA', 'RLT', 'RMU', 'RN', 'RPT', 
        'RT', 'SAF', 'SBM', 'SCD', 'SCH', 'SCM', 'SDS', 'SFC', 'SFD', 'SLD', 'SLL', 'SOW', 'SPC', 'SPL', 'SPR', 'SRV', 
        'STQ', 'SYM', 'TBE', 'TBS', 'TBT', 'TDS', 'TEQ', 'TMN', 'TN', 'UFD', 'UID', 'UT', 'VDB', 'VE', 'VIB', 'VIS', 
        'VMC', 'VMP', 'VO', 'VPQ', 'WBS', 'WIS', 'WPK', 'WPQ', 'WPR', 'WPS', 'WRR'
    ]
    seq_values = [f'{i:04d}' for i in range(5000)]  # From 0000 until 4999
    no_sheet_values = [f'{i:03d}' for i in range(500)]  # From 000 until 499
    rev_values = [f'{chr(65 + i)}{j}' for i in range(26) for j in range(1, 11)]  # A1 to Z10
    year_values = [str(year) for year in range(1970, 2024)]
    descp_values = ['AFB', 'AFC', 'AFDD', 'ASBUILT', 'CAN', 'FINAL', 'IDC', 'IFA', 'IFC', 'IFI', 'IFP', 'IFR', 'INF']

    # Write data to SourceSheet for each column vertically
    data_sets = [[''] * len(doc_types), doc_types, seq_values, no_sheet_values, [''] * len(doc_types), rev_values, year_values, descp_values]
      
    # Iterate through data_sets and update SourceSheet
    for col, data in enumerate(data_sets, start=1):
        source_sheet.cell(row=1, column=col).value = headers[col - 1]  # Writing header
        for idx, value in enumerate(data, start=1):
            source_sheet.cell(row=idx + 1, column=col).value = value
                
    # Apply data validation to MainSheet for specific columns except 'FileName' and 'Title'
    for col, header in enumerate(headers, start=1):
        if header != 'FileName' and header != 'Title':
            dv = DataValidation(type="list", formula1=f'=SourceSheet!${get_column_letter(col)}$2:${get_column_letter(col)}5002')
            main_sheet.add_data_validation(dv)
            for row in tqdm(range(2, len(pdf_filenames) + 2), desc=f"Applying {header} validation", unit="row"):
                main_sheet.cell(row=row, column=col).value = ''
                dv.add(main_sheet.cell(row=row, column=col))

    excel_path = os.path.join(folder_path, excel_filename)
    workbook.save(excel_path)
    print(f"Data saved to {excel_path}")

# Function to display a popup message
def display_popup_message(message):
    tk.messagebox.showinfo("Extraction Complete", message)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/every_single_file')
def every_single_file():
    return render_template('every_single_file.html')

# Function to handle data from 'Every Single File' form
@app.route('/process_every_single_file', methods=['POST'])
def process_every_single_file():
    folder_path = request.form['folder_path']
    pdf_paths = get_pdf_paths(folder_path)

    # Process each PDF and draw bounding boxes
    for pdf_path in pdf_paths:
        draw_bounding_boxes_single(pdf_path)

    # Save extracted data to an Excel file
    save_to_excel(pdf_filenames, extracted_texts, folder_path)

    # Display a popup message indicating extraction completion and where the files are saved
    message = f"All files have been successfully extracted and saved in the folder:\n{folder_path}"
    display_popup_message(message)

    # Once processing is done, set the success message
    success_message = "Data Successfully Extracted"

    return render_template('first_file.html', success_message=success_message)

@app.route('/first_file')
def first_file():
    return render_template('first_file.html')

# Function to handle data from 'Every Single File' form
@app.route('/process_first_file', methods=['POST'])
def process_first_file():
    folder_path = request.form['folder_path']
    pdf_paths = get_pdf_paths(folder_path)

    # Process the first PDF and draw bounding boxes
    if pdf_paths:
        draw_bounding_boxes_first(pdf_paths[0])
        bounding_boxes_drawn = True

    # Use the bounding boxes from the first PDF for the rest of the PDFs
    for pdf_path in pdf_paths[1:]:
        pdf_filenames.append(os.path.basename(pdf_path))
        extracted_data = extract_data(pdf_path, bounding_boxes)
        extracted_texts.append(extracted_data)

    # Save extracted data to an Excel file
    save_to_excel(pdf_filenames, extracted_texts, folder_path)

    # Display a popup message indicating extraction completion and where the files are saved
    message = f"All files have been successfully extracted and saved in the folder:\n{folder_path}"
    display_popup_message(message)

    # Once processing is done, set the success message
    success_message = "Data Successfully Extracted"

    return render_template('first_file.html', success_message=success_message)

if __name__ == '__main__':
    app.run(debug=True)
